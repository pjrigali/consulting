"""
Functions for reading in different file types.
"""
import csv
import zipfile
from xml.etree.cElementTree import XML
from pathlib import Path
from openpyxl import load_workbook


# Read csv's.
def read_csv(path: str, nrows: int = 0) -> list:
    """Simple csv reader, fixes keys, returns the csv as a list of dicts."""
    # Check if file is a csv.
    if not path.endswith('.csv'):
        path = path + '.csv'

    with open(path, mode='r', encoding='utf-8-sig') as file:
        csv_reader = csv.DictReader(file)

        # Catch and fix key names.
        if csv_reader.fieldnames:
            csv_reader.fieldnames = [name.lstrip("\ufeff") for name in csv_reader.fieldnames]

        # Only read in x lines.
        if nrows:
            return [next(csv_reader) for _ in range(nrows)]
        else:
            return [row for row in csv_reader]


# Update file version. Useful for creating successive csv outputs.
def update_file_version(folder_path: str, file_name: str) -> str:
    """
    Takes an input folder and file name, returns the next file version. This allow different versioning of files.
    Nest this in the save_csv(file_path=update_file_version(folder_path='your folder', file_name='your filenanme'), data).
    """
    root = Path(folder_path) if folder_path else Path.cwd()
    n = 1
    while True:
        n_filename = root / f'{file_name}_v{n}.csv'
        if not n_filename.exists():
            return str(n_filename)
        n += 1

# Old attempt for version updating.
# import os
# import re

# # New function to handle versioning and saving the data
# def update_file_version(folder_path: str, base_filename: str) -> str:
#     # Get list of all files in the folder
#     if folder_path:
#         files = os.listdir(folder_path)
#     else:
#         files = os.listdir()
   
#     # Filter out the CSV files
#     csv_files = [file for file in files if file.endswith('.csv') and file.startswith(base_filename)]
#     pattern = re.compile(rf"^{re.escape(base_filename)}_v(\d+)\.csv$")
   
#     version_numbers = []
#     for file in csv_files:
#         match = pattern.match(file)
#         if match:
#             version_numbers.append(int(match.group(1)))
   
#     # Determine the next version number
#     next_version = max(version_numbers, default=0) + 1

#     return os.path.join(folder_path, f"{base_filename}_v{next_version}.csv")


# Save a list of dictionaries to a csv.
def save_csv(file_path: str, data: list, sort_columns: bool = False) -> None:
    """Takes a list of dictionaries and saves them to a csv. Assumes all the dict's have the same keys."""
    with open(file_path, mode='w', encoding='utf-8-sig', newline='') as file:
        # I know, first row? gross.
        cols = data[0].keys()
    
        if sort_columns:
            cols.sort()
        
        csv_writer = csv.DictWriter(file, fieldnames=cols, restval='', extrasaction='ignore')        
        csv_writer.writeheader()
        csv_writer.writerows(data)

    print(f'File created: ({file_path})')
    return None


# Read in a word doc file.
def read_docx(file_path: str) -> str:
    """Takes a file location/name and returns the contents of a word document as str."""
    # Capture the file.
    wn = '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}'
    p, t = wn + 'p', wn + 't'
    document = zipfile.ZipFile(file_path)
    xml_content = document.read('word/document.xml')
    document.close()
    tree = XML(xml_content)

    # Read in the paragraphs.
    paragraphs = []
    for paragraph in tree.iter(p):
        texts = [node.text for node in paragraph.iter(t) if node.text]
        if texts:
            paragraphs.append(''.join(texts))
    return '\n'.join(paragraphs)


# Read in an excel file.
def read_excel(file_path: str, sheet_name: str) -> list:
    """Reads an xlsx file. Returns a list (workbook) of list (sheets) of dictionaries (rows)."""
    # Cheating, open the file.
    wb = load_workbook(filename=file_path, data_only=True, read_only=True)

    # Validate sheet is present, grap list of sheets present.
    sheets = []
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            try:
                wb.close()
            except Exception:
                pass
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
        else:
            sheets.append(sheet_name)
    else:
        sheets = wb.sheetnames

    # Capture the sheets
    results = []
    for sheet in sheets:
        # Grab rows.
        rows_iter = wb[sheet].iter_rows(values_only=True)

        # Check if the sheet has data, else skip.
        try:
            raw_headers = next(rows_iter)
        except StopIteration:
            results.append([{}])
            continue

        # Normalize headers: strings, trimmed; handle None/empty and duplicates
        seen_counts = {}
        headers = []
        for h in raw_headers:
            base = (str(h).strip() if h is not None else "").strip()
            if not base:
                base = "column"
            if base in seen_counts:
                seen_counts[base] += 1
                norm = f"{base}_{seen_counts[base]}"
            else:
                seen_counts[base] = 0
                norm = base
            headers.append(norm)

        # Parse the sheet.
        result = []
        for row in rows_iter:
            row_dict = {i: None for i in headers}
            if row is None:
                result.append(row_dict)
            else:
                # Would rather just dict(zip()) but whatever.
                for idx, key in enumerate(headers):
                    row_dict[key] = row[idx] if idx < len(row) else None
                result.append(row_dict)
        results.append(result)
    wb.close()
    return results
